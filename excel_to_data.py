import json
import re
import openpyxl

EXCEL_FILE = "12 Week Upper Lower Program v5 Updated.xlsx"
DATA_JS    = "data.js"

DAY_META = {
    "Mon": {"day": "Monday",    "label": "Upper Day 1"},
    "Tue": {"day": "Tuesday",   "label": "Lower Day 1"},
    "Wed": {"day": "Wednesday", "label": "Accessory"},
    "Thu": {"day": "Thursday",  "label": "Upper Day 2"},
    "Fri": {"day": "Friday",    "label": "Lower Day 2"},
}

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri"]


def val(v):
    s = str(v).strip() if v is not None else ""
    return "" if s == "None" else s


def parse_exercise(row):
    block    = val(row[0])
    exercise = val(row[1])
    sets     = val(row[2])
    reps     = val(row[3])
    tempo    = val(row[4])
    rest     = val(row[5])
    notes    = val(row[6])

    obj = {
        "block":    block,
        "exercise": exercise,
        "sets":     sets,
        "reps":     reps,
        "tempo":    tempo,
        "rest":     rest,
        "notes":    notes,
    }

    if block.startswith("FINISHER"):
        obj["isFinisher"] = True
    elif block.startswith("IWT"):
        obj["isBonus"]    = False
        obj["isFinisher"] = False
    elif "BONUS" in block:
        obj["isBonus"] = True
    else:
        obj["isBonus"] = False

    return obj


def load_tracker():
    with open(DATA_JS, "r", encoding="utf-8") as f:
        raw = f.read()
    json_str = re.sub(r"^const PROGRAM\s*=\s*", "", raw.strip()).rstrip(";").rstrip()
    return json.loads(json_str)["tracker"]


def build_program():
    tracker = load_tracker()

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # Collect sheets matching W\d\d (Mon|Tue|Wed|Thu|Fri)
    sheet_pat = re.compile(r"^W(\d{2})\s+(Mon|Tue|Wed|Thu|Fri)$")
    weeks_dict = {}

    for name in wb.sheetnames:
        m = sheet_pat.match(name)
        if not m:
            continue
        week_num = int(m.group(1))
        abbr     = m.group(2)
        ws       = wb[name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        title = val(rows[0][0])

        exercises = []
        for row in rows[2:]:  # skip header row
            if not any(v is not None for v in row):
                continue
            exercises.append(parse_exercise(row))

        day_obj = {
            "dayAbbr":   abbr,
            "day":       DAY_META[abbr]["day"],
            "label":     DAY_META[abbr]["label"],
            "title":     title,
            "exercises": exercises,
        }

        weeks_dict.setdefault(week_num, {})[abbr] = day_obj

    # Build ordered weeks array
    weeks = []
    for wnum in sorted(weeks_dict.keys()):
        days = [weeks_dict[wnum][d] for d in DAY_ORDER if d in weeks_dict[wnum]]
        weeks.append({"week": wnum, "days": days})

    return {"weeks": weeks, "tracker": tracker}


def main():
    program = build_program()
    with open(DATA_JS, "w", encoding="utf-8") as f:
        f.write("const PROGRAM = ")
        json.dump(program, f, indent=2, ensure_ascii=False)
        f.write(";\n")
    total_exercises = sum(
        len(day["exercises"])
        for week in program["weeks"]
        for day in week["days"]
    )
    print(f"Done — {len(program['weeks'])} weeks, {total_exercises} exercises written to {DATA_JS}.")


if __name__ == "__main__":
    main()
